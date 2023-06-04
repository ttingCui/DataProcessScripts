# -*- coding = utf-8 -*-
# 2023/5/15 16:09
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# url = 'https://www.paperdigest.org/2023/04/most-influential-acl-papers-2023-04/'  # 替换为目标网站的URL
# url = 'https://www.paperdigest.org/2023/04/most-influential-emnlp-papers-2023-04/'  # 替换为目标网站的URL
url = 'https://www.paperdigest.org/2023/04/most-influential-naacl-papers-2023-04/'  # 替换为目标网站的URL

# 发送HTTP请求获取网页内容
response = requests.get(url)
response.encoding = 'utf-8'  # 指定编码方式为 UTF-8

# 创建BeautifulSoup对象并解析网页内容
soup = BeautifulSoup(response.content, 'html.parser')

# 使用合适的选择器选择包含论文题目的HTML元素
# 获取包含论文标题的所有<tr>元素
rows = soup.select('table tr')

# 创建Excel工作簿
workbook = Workbook()
sheet = workbook.active

# 遍历所有<tr>元素，并提取其中第三个<td>元素中的<a>标签
for row_index, row in enumerate(rows, start=1):
    columns = row.select('td')
    if len(columns) >= 3:
        title_year = columns[0].text.strip()
        # title_month = columns[1].text.strip()
        title_id = columns[2].select_one('a')
        title_link = columns[2].select_one('a b')
        title_id = title_id['href']
        # if title_link:
        title_name = title_link.text.strip()
        # 将内容写入Excel表格中的对应单元格
        sheet.cell(row=row_index, column=1).value = title_year
        sheet.cell(row=row_index, column=2).value = title_name
        sheet.cell(row=row_index, column=3).value = title_id

    # 保存Excel文件
    workbook.save('output_naacl.xlsx')

